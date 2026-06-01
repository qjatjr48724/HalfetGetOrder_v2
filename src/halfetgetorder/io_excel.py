import os
import sys
import json
import openpyxl
from .godo import fetch_goods_base_specs
from datetime import date
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.worksheet import Worksheet

from .utils import visual_len, _to_int, _to_float
from .utils import _fmt_dt, get_box_count_from_items

header_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
center = Alignment(horizontal='center', vertical='center', wrap_text=False)
thin = Side(style="thin", color="000000")
thick = Side(style="thick", color="000000")


def get_project_root() -> str:
    """
    프로젝트 루트 경로를 반환.

    - 소스에서 실행할 때:
        io_excel.py 기준으로 ../../ 올라간 폴더 (HalfetGetOrder)
    - PyInstaller exe로 실행할 때:
        exe가 위치한 폴더 (dist) 기준
    """
    # PyInstaller로 빌드된 실행 파일 여부
    if getattr(sys, "frozen", False):
        # exe가 있는 폴더
        exe_dir = os.path.dirname(sys.executable)
        return exe_dir

    # 일반 파이썬 실행일 때 (python -m halfetgetorder)
    base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.abspath(os.path.join(base_dir, "..", ".."))


# ─────────────────────────────────────────────────────────
# Rich Text(한 셀 안에 서로 다른 스타일) 지원 여부 체크
# ─────────────────────────────────────────────────────────
try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    RICH_TEXT_AVAILABLE = True
except ImportError:
    RICH_TEXT_AVAILABLE = False


# ─────────────────────────────────────────────────────────
# 주문수집 엑셀 (주문내역)
# ─────────────────────────────────────────────────────────
def create_orders_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주문내역"
    # C열(총 상품결제금액)과 D열(체크) 사이에 '체크' 열 추가
    headers = [
        '플랫폼',           # A
        '주문일시',         # B
        '총 상품결제금액',   # C
        '체크',             # D
        '수취인 이름',      # E
        '상품명 + 옵션명',  # F
        '수량',             # G
        '등록옵션명',       # H
        '출고예정일',       # I  ← 새로 추가
        '배송메세지',       # J
    ]

    ws.append(headers)
    for c in ws[1]:
        c.fill = header_fill
    return wb, ws


def create_orders_workbook(
    coupang_orders: list[dict],
    godo_grouped_orders: list[dict],
) -> tuple[openpyxl.Workbook, Worksheet]:
    """
    주문수집(주문내역) 엑셀을 한 번에 만드는 헬퍼.

    - coupang_orders: coupang.normalize_coupang_orders(...) 결과 리스트
    - godo_grouped_orders: godo.group_orders(...) 결과 리스트

    JSON 파일을 건드리지 않고, 이미 메모리에 올라온 API 응답만 사용.
    """
    wb, ws = create_orders_sheet()

    if coupang_orders:
        append_coupang_block(ws, coupang_orders)
    if godo_grouped_orders:
        append_godo_sets(ws, godo_grouped_orders)

    finalize_orders_sheet(ws)
    return wb, ws


def apply_border_block(ws, start_row, end_row, start_col=1, end_col=10):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )


def apply_thick_bottom(ws, block_start, block_end, start_col=1, end_col=10):
    if block_end < block_start:
        return
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=block_end, column=c)
        cell.border = Border(
            left=cell.border.left or thin,
            right=cell.border.right or thin,
            top=cell.border.top or thin,
            bottom=thick
        )
    # 굵은 테두리 시작 기준 컬럼도 수취인 이름(E열=5번)로 변경
    top_left = ws.cell(row=block_start, column=5)
    top_left.border = Border(
        left=top_left.border.left or thin,
        right=top_left.border.right or thin,
        top=top_left.border.top or thin,
        bottom=thick
    )


def merge_receiver_name(ws, start_row, end_row, name: str | None = None):
    # 수취인 이름이 이제 5열(E)이므로 5번 컬럼 기준으로 병합
    if end_row < start_row:
        return
    top = ws.cell(row=start_row, column=5)
    if name and str(name).strip():
        top.value = str(name).strip()
    elif not (top.value and str(top.value).strip()):
        for r in range(start_row + 1, end_row + 1):
            val = ws.cell(row=r, column=5).value
            if val and str(val).strip():
                top.value = str(val).strip()
                break
    if end_row > start_row:
        ws.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)
        top.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )


def finalize_orders_sheet(ws):
    ws.sheet_view.zoomScale = 75
    min_widths = {
        '플랫폼': 8,
        '주문일시': 16,
        '총 상품결제금액': 14,
        '체크': 6,
        '수취인 이름': 15,       # E열 15
        '상품명 + 옵션명': 65,   # F열 65
        '수량': 5,               # G열 5
        '등록옵션명': 25,        # H열 25
        '출고예정일': 16,        # I열 16 (신규)
        '배송메세지': 35         # J열 35
    }

    headers = [cell.value for cell in ws[1]]

    for col in ws.columns:
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        header = headers[col_idx - 1] if col_idx - 1 < len(headers) else ''

        max_len = visual_len(header)
        for cell in col:
            vlen = visual_len(cell.value)
            if vlen > max_len:
                max_len = vlen

            # 줄바꿈 설정
            if header in ('등록옵션명', '배송메세지'):
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            elif header == '상품명 + 옵션명' and vlen > 50:
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True
                )
            else:
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=False
                )

            if header == '등록옵션명':
                cell.number_format = '@'

        # ── 여기부터 열 너비 계산 로직 정리 ──
        auto_width = int(max_len * 0.5)
        if header == '등록옵션명':
            auto_width = int(max_len * 0.5) + 4

        if header == '상품명 + 옵션명':
            # 항상 65 고정
            target_width = 65
        elif header == '배송메세지':
            # 항상 35 고정
            target_width = 35
        else:
            target_width = max(auto_width, min_widths.get(header, 12))

        ws.column_dimensions[col_letter].width = target_width




    # 상품명+옵션명 열(F), 배송메세지 열(J) 기준으로 행 높이 조정
    for r in range(2, ws.max_row + 1):
        prod_cell = ws.cell(row=r, column=6)   # F열
        memo_cell = ws.cell(row=r, column=10)  # J열


        pclen = visual_len(prod_cell.value)
        mlen = visual_len(memo_cell.value)

        base_len = max(pclen, mlen)
        rd = ws.row_dimensions[r]

        # 이미 다른 데서 높이를 지정한 행(예: 부모행 height=65)은 건드리지 않는다
        if rd.height is not None:
            continue

        if base_len > 40:
            rd.height = 34
        else:
            rd.height = 24


    # 체크 열(D열에 값이 있고, F열이 '+ '로 시작하지 않는 = 부모행만 색상 변경)
    last_row = ws.max_row
    if last_row >= 2:
        fill_checked = PatternFill(
            start_color="FFE6FFCC",
            end_color="FFE6FFCC",
            fill_type="solid"
        )

        rule = FormulaRule(
            formula=['AND(LEN($D2)>0, LEFT($F2,2)<>" + ")'],
            fill=fill_checked
        )

        # A2 ~ J{마지막 행}까지 적용
        ws.conditional_formatting.add(f"A2:J{last_row}", rule)


    # 전체 글꼴 크기를 12로 통일 (기존 bold/italic, 색상 등은 유지)
    for row in ws.iter_rows():
        for cell in row:
            font = cell.font or Font()
            if font.size == 12:
                continue
            cell.font = font.copy(size=12)




def append_coupang_block(ws, coupang_orders):
    current_row = ws.max_row + 1
    for od in coupang_orders:
        block_start = current_row
        ordered_at = _fmt_dt(od.get('orderedAt') or od.get('orderDate', ''))

        total_price = 0.0
        for item in od.get('orderItems', []):
            qty = _to_int(item.get('shippingCount', 1), 1)
            sales_price = _to_float(item.get("salesPrice", 0))
            order_price = _to_float(item.get("orderPrice", 0))
            price = _to_float(item.get("price", 0))

            # 쿠팡 응답 기준:
            # - salesPrice: 단가
            # - orderPrice: 수량이 반영된 라인 합계인 케이스가 많음
            # 따라서 우선 salesPrice * qty 로 계산하되,
            # salesPrice 가 없으면 orderPrice(합계) → price * qty 순서로 fallback.
            if sales_price > 0:
                total_price += sales_price * qty
            elif order_price > 0:
                total_price += order_price
            else:
                total_price += price * qty
        total_price_str = f"{int(total_price):,}원"

        receiver_name = (
            (od.get('shippingAddress') or {}).get('name', '') or
            (od.get('receiver') or {}).get('name', '')
        )

        item_names = []
        total_qty = 0
        items = od.get('orderItems', []) or []

        for item in items:
            name = (
                item.get('sellerProductName')
                or item.get('vendorItemName')
                or item.get('productName')
                or ""
            )
            option = (
                item.get('sellerProductItemName')
                or item.get('vendorItemName')
                or ""
            )
            qty = _to_int(item.get('shippingCount', 1), 1)
            total_qty += qty
            if name and option and option != name:
                item_names.append(f"{name} / {option}")
            else:
                item_names.append(name or option)
        product_info = " / ".join([x for x in item_names if x])
        total_qty = total_qty or 1

        # 등록옵션명 (쿠팡 기준: 상품명 사용)
        reg_option_name = ""
        if items:
            first_item = items[0] or {}
            reg_option_name = (
                first_item.get('sellerProductName')
                or first_item.get('vendorItemName')
                or first_item.get('productName')
                or ""
            )

        # 출고예정일 (orderItems[*].estimatedShippingDate 중 가장 이른 날짜 사용)
        est_dates = [
            (it.get("estimatedShippingDate") or "").strip()
            for it in items
            if (it.get("estimatedShippingDate") or "").strip()
        ]
        est_shipping_display = ""
        if est_dates:
            # ISO 날짜 문자열 기준으로 가장 빠른 날짜 선택
            est_min = min(est_dates)
            est_shipping_display = _fmt_dt(est_min)


        # 쿠팡 배송메세지: parcelPrintMessage
        coupang_memo = od.get('parcelPrintMessage', '') or ''

        # A:플랫폼, B:주문일시, C:총금액, D:체크, E:수취인, F:상품+옵션, G:수량,
        # H:등록옵션명, I:출고예정일, J:배송메세지
        ws.append([
            "쿠팡",
            ordered_at,
            total_price_str,
            "",                   # 체크 열
            receiver_name,
            product_info,
            total_qty,
            reg_option_name,
            est_shipping_display,  # 새로 추가된 출고예정일
            coupang_memo,
        ])

        current_row += 1

        apply_border_block(ws, block_start, current_row - 1, 1, 10)
        merge_receiver_name(ws, block_start, current_row - 1)
        apply_thick_bottom(ws, block_start, current_row - 1, 1, 10)



# ─────────────────────────────────────────────────────────
# 고도몰 상품 기본 사양 / 옵션 관련
# ─────────────────────────────────────────────────────────
def _parse_short_desc_to_specs(short_desc: str) -> tuple[str, str]:
    """
    shortDescription 예시:
      'DeLL Latitude 5501 / Intel® Core™ i7-9850H / NVIDIA GeForce MX150 /
       NVMe SSD 512G / DDR4 32G / FHD ... / 윈도우11'

    '/' 로 나눈 뒤:
      index 3 → SSD 파트 (예: 'NVMe SSD 512G')
      index 4 → RAM 파트 (예: 'DDR4 32G')
    """
    if not short_desc:
        return "", ""

    parts = [p.strip() for p in str(short_desc).split("/") if p.strip()]

    ssd = parts[3].strip() if len(parts) > 3 else ""
    ram = parts[4].strip() if len(parts) > 4 else ""

    # (RAM, SSD) 순서로 반환
    return ram, ssd


def _build_base_specs_from_raw(raw) -> dict:
    """
    raw 를 {상품번호: {ram, ssd}} 형태로 정규화.

    지원 형태:
      1) 딕셔너리:
         {
           "1000001": { "ram": "16G", "ssd": "512G" }
           "1000002": { "shortDescription": "..." }
           "1000003": "DeLL Latitude 5501 / ... / NVMe SSD 512G / DDR4 32G / ..."
         }

      2) 리스트:
         [
           { "goodsNo": "1000001", "ram": "16G", "ssd": "512G" },
           { "goodsNo": "1000002", "shortDescription": "..." },
           { "goodsNo": "1000003", "shortDescription": "..." },
         ]
    """
    base_specs: dict[str, dict[str, str]] = {}

    # case 1: dict
    if isinstance(raw, dict):
        for key, val in raw.items():
            goods_key = str(key).strip()
            if not goods_key:
                continue

            ram = ""
            ssd = ""

            if isinstance(val, dict):
                ram = str(val.get("ram", "")).strip()
                ssd = str(val.get("ssd", "")).strip()
                short_desc = str(val.get("shortDescription", "")).strip()

                # ram/ssd 없으면 shortDescription에서 뽑기
                if short_desc and (not ram or not ssd):
                    ram2, ssd2 = _parse_short_desc_to_specs(short_desc)
                    ram = ram or ram2
                    ssd = ssd or ssd2
            else:
                # 값이 그냥 shortDescription 문자열인 경우
                short_desc = str(val).strip()
                if short_desc:
                    ram, ssd = _parse_short_desc_to_specs(short_desc)

            base_specs[goods_key] = {"ram": ram, "ssd": ssd}

    # case 2: list
    elif isinstance(raw, list):
        for row in raw:
            if not isinstance(row, dict):
                continue

            goods_key = str(
                row.get("goodsNo") or row.get("goodsCd") or ""
            ).strip()
            if not goods_key:
                continue

            ram = str(row.get("ram", "")).strip()
            ssd = str(row.get("ssd", "")).strip()
            short_desc = str(row.get("shortDescription", "")).strip()

            if short_desc and (not ram or not ssd):
                ram2, ssd2 = _parse_short_desc_to_specs(short_desc)
                ram = ram or ram2
                ssd = ssd or ssd2

            base_specs[goods_key] = {"ram": ram, "ssd": ssd}

    return base_specs


def load_godo_base_specs_map(path: str | None = None) -> dict:
    """
    고도몰 상품 기본 RAM/SSD 사양 로드.

    우선순위:
      1) 인자로 받은 path
      2) 프로젝트 루트의 godo_base_specs.json

    예전에는 godo_goods_all.json(상품 전체 캐시)을 함께 사용했지만,
    이제는 대용량 JSON 파일을 만들지 않고
    필요한 경우 Goods_Search API를 직접 호출하는 방식으로 변경했다.
    """
    project_root = get_project_root()

    candidates: list[str] = []
    if path:
        candidates.append(path)
    candidates.append(os.path.join(project_root, "godo_base_specs.json"))

    for p in candidates:
        if not p:
            continue
        if not os.path.exists(p):
            continue

        try:
            with open(p, "r", encoding="utf-8") as f:
                raw = json.load(f)
            specs = _build_base_specs_from_raw(raw)
        except Exception as e:
            print(f"⚠️ 기본 사양 파일({p})을 읽는 중 오류: {e}")
            continue

        if specs:
            print(f"[라벨] 고도몰 기본 RAM/SSD 사양 {len(specs)}건 로드 ({p})")
            return specs

    print("ℹ️ godo_base_specs.json 을 찾지 못했습니다. 고도몰 기본 RAM/SSD는 API에서 보완 조회합니다.")
    return {}



def get_godo_base_ram_ssd(parent: dict, base_specs_map: dict) -> tuple[str, str]:
    """
    고도몰 parent(본상품) 한 건에 대해 기본 RAM/SSD 를 조회.
    - 우선 goodsNo 로 찾고
    - 없으면 goodsCd 로도 한 번 더 찾아본다.
    """
    goods_no = str(parent.get("goodsNo") or "").strip()
    goods_cd = str(parent.get("goodsCd") or "").strip()

    spec = None
    if goods_no:
        spec = base_specs_map.get(goods_no)
    if spec is None and goods_cd:
        spec = base_specs_map.get(goods_cd)

    if not spec:
        return "", ""

    ram = str(spec.get("ram", "")).strip()
    ssd = str(spec.get("ssd", "")).strip()
    return ram, ssd



### 이제 사용하지 않는 함수
def load_godo_goods_map(path: str | None = None) -> dict:
    """
    goods_search로 미리 만들어둔 godo_goods_all.json 로드.
    key: goodsNo
    value: goods_search 응답 전체(dict)
    """
    # if path is None:
    #     project_root = get_project_root()
    #     path = os.path.join(project_root, "godo_goods_all.json")

    # if not os.path.exists(path):
    #     print("⚠️ godo_goods_all.json 파일을 찾을 수 없습니다. 기본 RAM/SSD는 비워둡니다.")
    #     return {}

    # with open(path, "r", encoding="utf-8") as f:
    #     return json.load(f)
    return {}


def get_base_specs_from_short_description(parent: dict, goods_map: dict) -> tuple[str, str]:
    """
    - 우선 parent(주문의 본상품) 안에 shortDescription 이 있으면 그걸 쓰고,
    - 없으면 godo_goods_all.json(goods_map)에서 goodsNo 로 찾아서 shortDescription을 가져온다.

    shortDescription 예시:
      DeLL Latitude 5501 / Intel® Core™ i7-9850H / NVIDIA GeForce MX150 /
      NVMe SSD 512G / DDR4 32G / FHD 1920×1080 해상도 (15.6인치) / 윈도우11

    / 로 split 한 후:
      0: 모델명
      1: CPU
      2: 그래픽
      3: SSD
      4: RAM
      5: 해상도
      6: 윈도우 버전

    여기서
      - 기본 SSD  → parts[3]
      - 기본 RAM  → parts[4]
    를 그대로 사용한다.
    """
    short_desc = (parent.get("shortDescription") or "").strip()

    goods_no = str(parent.get("goodsNo") or "").strip()

    # parent 안에 shortDescription 이 없으면 API로 한 번 더 조회
    if not short_desc and goods_no:
        try:
            short_desc = godo.fetch_goods_short_description(goods_no) or ""
        except Exception as e:
            print(f"⚠️ Goods_Search API에서 shortDescription 조회 실패 (goodsNo={goods_no}): {e}")
            short_desc = ""

    if not short_desc:
        return "", ""

    parts = [p.strip() for p in short_desc.split("/")]
    if len(parts) <= 4:
        return "", ""

    ssd_part = parts[3].strip()
    ram_part = parts[4].strip()

    return ram_part, ssd_part


def build_godo_option_text_from_children(children: list[dict]) -> str:
    """
    고도몰 parent 아래의 children(추가상품) 리스트에서
    추가상품명(goodsNm 계열)을 줄바꿈(\n)으로 이어서 옵션 문자열로 만든다.

    예:
      children = [
        {"goodsNm": "+ 메모리 8G→16G로 UP↑", ...},
        {"goodsNm": "+ 윈도우 복구 프로그램", ...},
      ]

      → "+ 메모리 8G→16G로 UP↑\n+ 윈도우 복구 프로그램"
    """
    names: list[str] = []

    for add in children or []:
        name = (
            add.get("goodsNm")
            or add.get("goodsNmStandard")
            or add.get("goodsNmView")
            or ""
        )
        name = str(name).strip()
        if name:
            names.append(name)

    return "\n".join(names)


# ─────────────────────────────────────────────────────────
# 쿠팡 라벨용 RAM/SSD/옵션 추출
# ─────────────────────────────────────────────────────────
def extract_specs_from_coupang_item(
    item: dict,
    keyskin_models: list[str] | None = None
):
    """
    쿠팡 orderItems[*]에서 RAM / SSD / 옵션 추출.
    - RAM: sellerProductItemName.split()[3]
    - SSD: sellerProductItemName.split()[2]
    - 옵션: [리브레, 원키] + (모델명에 키워드 포함되면 키스킨)
    """
    seller_item_name = item.get("sellerProductItemName") or ""
    tokens = seller_item_name.split()

    ram = ""
    ssd = ""

    if len(tokens) > 3:
        ram = tokens[3]
    if len(tokens) > 2:
        ssd = tokens[2]

    # 옵션 기본값
    options = ["리브레", "원키"]

    # 모델명 기반 키스킨 추가
    if keyskin_models:
        model_name = (
            item.get("sellerProductName")
            or item.get("vendorItemName")
            or item.get("productName")
            or ""
        )
        for kw in keyskin_models:
            if kw and kw in model_name:
                options.append("키스킨")
                break

    option_str = " / ".join(options)
    return ram, ssd, option_str


def append_godo_sets(ws, grouped_orders):
    """
    고도몰 주문을 엑셀 주문내역 시트에 추가.

    - 부모행(본상품)의 '상품명 + 옵션명' 셀(6열)에:
        상품명
        optionInfo(옵션명: 값 ...)
      이렇게 줄바꿈해서 표시.
    """
    # 자사몰 주문은 역순(최근 주문이 아래로)
    for grp in reversed(grouped_orders):
        if not grp.get("sets"):
            continue

        block_start = ws.max_row + 1
        first_parent = True
        receiver_name = (grp.get("receiver") or {}).get("name") or ""
        ordered_at = _fmt_dt(grp.get("orderedAt") or "")

        for s in grp["sets"]:
            p = s["parent"]
            goodsCd = (p.get('goodsCd') or p.get('goodsModel') or '').strip()
            goodsNm = (p.get('goodsNm') or p.get('goodsNmStandard') or '').strip()
            qty = _to_int(p.get('goodsCnt', 1), 1)
            price = _to_float(p.get('goodsPrice', 0.0), 0.0)

            # 상품명
            product_name = goodsNm or goodsCd

            # 1) orderoptionInfo / orderOptionInfo 우선
            option_info = (
                (p.get('orderoptionInfo') or '').strip()
                or (p.get('orderOptionInfo') or '').strip()
            )

            # 2) 없으면 optionInfo(JSON 문자열) 파싱
            if not option_info:
                raw_opt = (p.get('optionInfo') or '').strip()
                if raw_opt:
                    try:
                        opt_list = json.loads(raw_opt)  # [[옵션명, 옵션값, ...], ...]
                        parts = []
                        for opt in opt_list:
                            if isinstance(opt, (list, tuple)) and len(opt) >= 2:
                                name = str(opt[0]).strip()
                                val = str(opt[1]).strip()
                                if name and val:
                                    parts.append(f"{name}: {val}")
                        option_info = "\n".join(parts)
                    except Exception:
                        option_info = ""

            if option_info:
                product_info_parent = f"{product_name}\n{option_info}"
            else:
                product_info_parent = product_name

            reg_option_value = goodsCd

            # 세트 총 금액
            set_total = price * (qty or 1)
            for add in s["children"]:
                add_qty = _to_int(add.get('goodsCnt', 1), 1)
                add_price = _to_float(add.get('goodsPrice', 0.0), 0.0)
                set_total += add_price * add_qty
            total_price_str = f"{int(set_total):,}원"

            order_memo = (
                grp.get("orderMemo", "")
                or grp.get("orderInfo", {}).get("orderMemo", "")
            )

            ws.append([
                "고도몰",
                ordered_at if first_parent else "",
                total_price_str,
                "",   # 체크 열
                receiver_name if first_parent else "",
                product_info_parent,
                (qty or 1),
                reg_option_value,
                "",                         # 출고예정일 (고도몰은 비워둠)
                order_memo if first_parent else "",
            ])

            first_parent = False
            prow = ws.max_row

            # 부모 셀 스타일링 (6열)
            pcell = ws.cell(row=prow, column=6)

            if option_info and RICH_TEXT_AVAILABLE:
                pcell.value = CellRichText(
                    TextBlock(
                        text=product_name,
                        font=InlineFont(b=True)
                    ),
                    TextBlock(
                        text="\n" + option_info,
                        font=InlineFont(
                            i=True,
                            
                        )
                    ),
                )
            else:
                pcell.value = product_info_parent
                pcell.font = Font(bold=True)

            pcell.alignment = Alignment(
                horizontal='left',
                vertical='center',
                wrap_text=True
            )
            pcell.fill = PatternFill(
                start_color="FFF7F7F7",
                end_color="FFF7F7F7",
                fill_type="solid"
            )

            ws.row_dimensions[prow].height = 65

            # 자식(추가옵션) 행
            for add in s["children"]:
                add_name = (add.get('goodsNm') or add.get('goodsNmStandard') or '').strip()
                add_qty = _to_int(add.get('goodsCnt', 1), 1)
                ws.append(["", "", "", "", "", f"+ {add_name}", add_qty, "", "", ""])
                crow = ws.max_row
                ccell = ws.cell(row=crow, column=6)
                ccell.font = Font(italic=True)
                ccell.alignment = Alignment(
                    horizontal='left',
                    vertical='center',
                    indent=1
                )

        block_end = ws.max_row
        if block_end >= block_start:
            apply_border_block(ws, block_start, block_end, 1, 10)
            merge_receiver_name(ws, block_start, block_end, receiver_name)
            apply_thick_bottom(ws, block_start, block_end, 1, 10)




# ─────────────────────────────────────────────────────────
# 라벨 출력용 엑셀
# ─────────────────────────────────────────────────────────
def create_label_workbook(
    coupang_orders: list[dict],
    godo_grouped_orders: list[dict],
    godo_base_specs_path: str | None = None,
    godo_goods_all_path: str | None = None,
    godo_add_goods_map_path: str | None = None,  # ← 지금은 사용하지 않지만, 시그니처 유지
) -> tuple[openpyxl.Workbook, Worksheet]:
    """
    라벨 출력용 엑셀 워크북 생성.

    - 시트명: '라벨'
    - 열: [플랫폼, 이름, 모델명, 램, SSD, 옵션]

    쿠팡:
      - RAM/SSD/옵션: extract_specs_from_coupang_item()
      - shippingCount(수량) 만큼 행 반복

    고도몰:
      - 자사몰 주문은 역순(최근 주문이 아래로)
      - goodsCnt(수량) 만큼 행 반복
      - 모델명 셀: "모델명\\noptionInfo"
      - 옵션 셀: children[*].goodsNm 을 줄바꿈으로 표시
      - 기본 RAM/SSD: godo_base_specs_map + godo_goods_all.json
    """
    keyskin_models = [
        "그램 17",
        "Latitude 5520",
        "키스킨 포함",
        "키보드 키스킨",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "라벨"

    headers = ["플랫폼", "이름", "모델명", "램", "SSD", "옵션"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill(
            start_color="D8E4BC",
            end_color="D8E4BC",
            fill_type="solid",
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 고도몰 기본 사양/상품 정보 로드
    missing_base_spec_ids: set[str] = set()

    # 1) 쿠팡 라벨
    for od in coupang_orders:
        receiver_name = (
            (od.get("shippingAddress") or {}).get("name", "")
            or (od.get("receiver") or {}).get("name", "")
        )

        for item in od.get("orderItems", []):
            model_name = (
                item.get("sellerProductName")
                or item.get("vendorItemName")
                or item.get("productName")
                or ""
            )

            ram, ssd, option_str = extract_specs_from_coupang_item(
                item,
                keyskin_models=keyskin_models,
            )

            qty = _to_int(item.get("shippingCount", 1), 1)
            if qty <= 0:
                qty = 1

            for _ in range(qty):
                ws.append(
                    [
                        "쿠",
                        receiver_name,
                        model_name,
                        ram,
                        ssd,
                        option_str,
                    ]
                )

    # 2) 고도몰 라벨 (역순)
    for grp in reversed(godo_grouped_orders or []):
        receiver_name = grp.get("receiver", {}).get("name", "")

        for s in grp.get("sets", []):
            parent = s.get("parent", {})
            children = s.get("children", []) or []

            model_name = (parent.get("goodsCd") or "").strip()

            # ✅ 기본 RAM/SSD는 json 파일을 전혀 보지 않고, goodsNo 기준으로 API에서 바로 조회
            goods_no = str(parent.get("goodsNo") or "").strip()
            base_ram, base_ssd = fetch_goods_base_specs(goods_no)

            if not (base_ram or base_ssd):
                key = goods_no or model_name
                if key:
                    missing_base_spec_ids.add(key)

            # optionInfo 문자열 만들기
            option_info = (
                (parent.get("orderoptionInfo") or "").strip()
                or (parent.get("orderOptionInfo") or "").strip()
            )

            if not option_info:
                raw_opt = (parent.get("optionInfo") or "").strip()
                if raw_opt:
                    try:
                        opt_list = json.loads(raw_opt)
                        parts: list[str] = []
                        for opt in opt_list:
                            if isinstance(opt, (list, tuple)) and len(opt) >= 2:
                                name = str(opt[0]).strip()
                                val = str(opt[1]).strip()
                                if name and val:
                                    parts.append(f"{name}: {val}")
                        option_info = " / ".join(parts)
                    except Exception:
                        option_info = ""

            model_cell_value = model_name
            if option_info:
                model_cell_value = f"{model_name}\n{option_info}"

            option_str = build_godo_option_text_from_children(children)

            qty = _to_int(parent.get("goodsCnt", 1), 1)
            if qty <= 0:
                qty = 1

            for _ in range(qty):
                ws.append(
                    [
                        "자",
                        receiver_name,
                        model_cell_value,
                        base_ram,
                        base_ssd,
                        option_str,
                    ]
                )

    # 서식 설정
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter in ("C", "F"):
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

    width_map = {
        "A": 10,
        "B": 18,
        "C": 45,
        "D": 12,
        "E": 12,
        "F": 30,
    }
    for col, w in width_map.items():
        ws.column_dimensions[col].width = w

    ws.sheet_view.zoomScale = 90

    if missing_base_spec_ids:
        print(
            "[라벨] RAM/SSD 기본사양을 찾지 못한 고도몰 상품번호/코드: "
            + ", ".join(sorted(missing_base_spec_ids))
        )

    return wb, ws


# ─────────────────────────────────────────────────────────
# 대한통운 송장등록 엑셀
# ─────────────────────────────────────────────────────────
def create_waybill_workbook(coupang_orders):
    """
    대한통운 송장등록용 엑셀 워크북 생성.
    - 시트명: '판매 주문수집'
    - 열 구조: 기존 단일 파일 버전의 first_col1 과 동일
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "판매 주문수집"
    ws.sheet_view.zoomScale = 75

    header = [
        '예약구분', '집하예정일', '받는분성명', '받는분전화번호', '받는분기타연락처',
        '받는분우편번호', '받는분주소(전체, 분할)', '운송장번호', '고객주문번호',
        '품목명', '박스수량', '박스타입', '기본운임', '배송메세지1',
        '배송메세지2', '품목명', '운임구분'
    ]
    ws.append(header)

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    today_str = date.today().strftime('%Y%m%d')

    for od in coupang_orders:
        name = od.get("name", "")
        phone = od.get("phone", "")
        addr1 = od.get("addr1", "")
        addr2 = od.get("addr2", "")
        zipcode = od.get("zipcode", "")
        address = f"{addr1} {addr2}".strip()

        box_cnt = get_box_count_from_items(od.get("items", []))
        platform_name = "쿠팡"

        row = [
            "일반",
            today_str,
            name,
            phone,
            "",
            zipcode,
            address,
            "",
            "",
            "",
            box_cnt,
            "",
            "",
            "",
            platform_name,
            "",
            "신용"
        ]
        ws.append(row)

    center_align = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align

    for column_cells in ws.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max_len * 1.3 + 2

    return wb, ws
